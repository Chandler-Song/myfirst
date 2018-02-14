import openpyxl
from openpyxl.styles import Font
#N值
N = 20
#创建表格
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
#设置粗体
font1 = Font(bold = True)
for i in range(2, N + 2):
    sheet.cell(row=1, column=i).font = font1
    sheet.cell(row = 1, column = i).value = i - 1
for i in range(2, N + 2):
    sheet.cell(row=i, column=1).font = font1
    sheet.cell(row = i, column = 1).value = i - 1
for i in range(2, N + 2):
    for m in range(2, N + 2):
        sheet.cell(row = i, column = m).value = sheet.cell(row = i, column = 1).value * sheet.cell(row = 1,column = m).value
wb.save("ok.xlsx")