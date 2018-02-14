#sendDuesReminders.py - sends emails based on payment status in spreadsheet.

import openpyxl, smtplib, sys

#open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook("duesRecords.xlsx")
sheet = wb.get_active_sheet()
lastCol = sheet.max_column
latestMonth = sheet.cell(row = 1, column =  lastCol).value
#Check each menber's payment status.
unpaidMenbers = {}
for r  in range(2,sheet.max_row + 1):
    payment = sheet.cell(row = r, column = lastCol).value
    if payment != "paid":
        name = sheet.cell(row = r, column = 1).value
        email = sheet.cell(row = r, column = 2).value
        unpaidMenbers[name] = email
#log in to email account
smptObj = smtplib.SMTP_SSL("smtp.163.com", 465)
smptObj.ehlo()
smptObj.login("liubiyongge@163.com", "13707219489lby")

#Send out reminder emails.
for name, email in unpaidMenbers.items():
    body = "Subject: %s dues unpaid.\nDear %s,\nRecord show that you have not paid dues for %s. Please make this payment as soon as possible. Thank you!"%(latestMonth, name, latestMonth)
    print("Send email to %s..." % email)
    sendMailStatus = smptObj.sendmail("liubiyongge@163.com", email, body)
    if sendMailStatus != {}:
        print("There was a problem sending email to %s: %s" %(email, sendMailStatus))
smptObj.quit()
